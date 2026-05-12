"""
Excel ingestion + validation for the onboarding employee import.

All parsing is openpyxl-based. The same parser powers both code paths off
the chat endpoint:
  - intent=employees_upload  (user-supplied file)
  - intent=employees_demo    (Demo_Data_Employee_Worklynx.xlsx)

Validation surfaces every issue the UI's ValidationReport renders:
missing_columns, invalid_rows, duplicate_entries, empty_required_fields,
invalid_date_formats, warnings.
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from .utils import (
    DATE_FIELDS,
    DEMO_EMPLOYEES_XLSX,
    NON_EMPTY_FIELDS,
    REQUIRED_COLUMNS as DEFAULT_REQUIRED_COLUMNS,
    TEMPLATE_EMPLOYEES_XLSX,
    cell_value,
    is_valid_email,
    to_iso_date,
)


DATA_SHEET_CANDIDATES = ("Data", "Employees", "Sheet1")


def _open_workbook(source: BinaryIO | Path | str):
    return load_workbook(filename=source, data_only=True, read_only=False)


def _pick_data_sheet(wb):
    for name in DATA_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    # Fall back to the first sheet
    return wb[wb.sheetnames[0]]


def _header_row(ws) -> list[str]:
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    return [(h or "").strip() if isinstance(h, str) else (str(h).strip() if h is not None else "") for h in first]


def _build_row_dict(headers: list[str], row: tuple) -> dict[str, Any]:
    return {headers[i]: row[i] for i in range(min(len(headers), len(row)))}


def get_required_columns() -> list[str]:
    """Read the canonical column list directly from the live template file.

    This makes the validator self-adjusting: if the user edits the template
    (e.g. removes the `_id` column), uploads validate against the new shape
    automatically. Falls back to the hard-coded default if the template is
    missing or unreadable.
    """
    if not TEMPLATE_EMPLOYEES_XLSX.exists():
        return list(DEFAULT_required_columns)
    try:
        wb = load_workbook(TEMPLATE_EMPLOYEES_XLSX, data_only=True, read_only=True)
        ws = _pick_data_sheet(wb)
        headers = [h for h in _header_row(ws) if h]
        wb.close()
        return headers or list(DEFAULT_required_columns)
    except Exception:
        return list(DEFAULT_required_columns)


def parse_employees(source: BinaryIO | Path | str) -> dict[str, Any]:
    """
    Parse an employee xlsx and produce both the structured employee list
    and a full validation breakdown.

    Returns:
        {
            "employees": [...],
            "validation": {
                "missing_columns": [...],
                "invalid_rows": [...],
                "duplicate_entries": [...],
                "empty_required_fields": [...],
                "invalid_date_formats": [...],
                "warnings": [...],
            },
            "row_count": int,
        }
    """
    wb = _open_workbook(source)
    ws = _pick_data_sheet(wb)
    headers = _header_row(ws)

    # Source of truth: the live template's header row. Falls back to the
    # hard-coded default if the template is missing.
    required_columns = get_required_columns()
    missing_columns = [c for c in required_columns if c not in headers]
    extra_columns = [h for h in headers if h and h not in required_columns]

    employees: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, Any]] = []
    duplicate_entries: list[dict[str, Any]] = []
    empty_required_fields: list[dict[str, Any]] = []
    invalid_date_formats: list[dict[str, Any]] = []
    warnings: list[str] = []

    seen_emails: dict[str, int] = {}
    seen_codes: dict[str, int] = {}

    # Begin from row 2 (1-indexed). The user-facing row number includes the
    # header row so it matches what they see in Excel.
    for excel_row_num, raw in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(v in (None, "") for v in raw):
            continue

        row = _build_row_dict(headers, raw)

        # Empty required fields
        for field in NON_EMPTY_FIELDS:
            if field in headers:
                if row.get(field) in (None, "") or (
                    isinstance(row.get(field), str) and not row.get(field).strip()
                ):
                    empty_required_fields.append({"row": excel_row_num, "field": field})

        # Email format
        email_raw = row.get("Email")
        if email_raw and isinstance(email_raw, str) and not is_valid_email(email_raw):
            invalid_rows.append({
                "row": excel_row_num,
                "reason": f"Invalid email format: {email_raw}",
            })

        # Date validation + normalisation
        for date_field in DATE_FIELDS:
            if date_field in headers:
                value = row.get(date_field)
                if value in (None, ""):
                    continue
                iso = to_iso_date(value)
                if iso is None:
                    invalid_date_formats.append({
                        "row": excel_row_num,
                        "field": date_field,
                        "value": str(value),
                    })
                else:
                    row[date_field] = iso

        # Duplicate detection
        email_key = (row.get("Email") or "").strip().lower() if isinstance(row.get("Email"), str) else ""
        code_key = (row.get("Employee Code") or "").strip() if isinstance(row.get("Employee Code"), str) else ""

        if email_key:
            if email_key in seen_emails:
                duplicate_entries.append({
                    "row": excel_row_num,
                    "field": "Email",
                    "value": email_key,
                    "first_seen_row": seen_emails[email_key],
                })
            else:
                seen_emails[email_key] = excel_row_num

        if code_key:
            if code_key in seen_codes:
                duplicate_entries.append({
                    "row": excel_row_num,
                    "field": "Employee Code",
                    "value": code_key,
                    "first_seen_row": seen_codes[code_key],
                })
            else:
                seen_codes[code_key] = excel_row_num

        # Build serialisable employee record
        employee = {h: cell_value(row.get(h)) for h in headers if h in required_columns}
        # Preserve any extra columns the file may carry, but namespace them so
        # the schema-aware UI can ignore them safely.
        extras = {h: cell_value(row.get(h)) for h in headers if h and h not in required_columns}
        if extras:
            employee["_extras"] = extras

        employees.append(employee)

    if missing_columns:
        warnings.append(
            f"{len(missing_columns)} required column(s) missing — fields will be blank in import."
        )
    if extra_columns:
        warnings.append(
            f"{len(extra_columns)} unexpected column(s) found "
            f"({', '.join(extra_columns[:4])}{'…' if len(extra_columns) > 4 else ''}) "
            "— they will be ignored."
        )
    if not employees:
        warnings.append("No data rows detected. Did you fill out the Data sheet?")

    return {
        "employees": employees,
        "row_count": len(employees),
        "validation": {
            "missing_columns": missing_columns,
            "invalid_rows": invalid_rows,
            "duplicate_entries": duplicate_entries,
            "empty_required_fields": empty_required_fields,
            "invalid_date_formats": invalid_date_formats,
            "warnings": warnings,
        },
    }


def parse_demo_employees() -> dict[str, Any]:
    """
    Read the project-root demo file and patch every row so it is always
    safe to import:

      - Missing / invalid `Date of Joining` → random date in the last 5 years
      - Missing / invalid `Date of Birth`   → random date producing an adult
        between 25 and 55 years old
      - Empty `Gender`                       → random pick
      - Other empty required fields          → sensible default

    Returns the same shape as `parse_employees`, but with no
    `empty_required_fields` / `invalid_date_formats` reported on demo data —
    the user explicitly opted into the canned demo.
    """
    if not DEMO_EMPLOYEES_XLSX.exists():
        return {
            "employees": [],
            "row_count": 0,
            "validation": {
                "missing_columns": [],
                "invalid_rows": [],
                "duplicate_entries": [],
                "empty_required_fields": [],
                "invalid_date_formats": [],
                "warnings": [f"Demo file not found at {DEMO_EMPLOYEES_XLSX}"],
            },
        }

    result = parse_employees(DEMO_EMPLOYEES_XLSX)
    result["employees"] = [_sanitize_demo_row(emp) for emp in result["employees"]]

    # Demo data is curated — clear the noisy validation buckets so the user
    # sees a clean table without bogus warnings.
    v = result["validation"]
    v["empty_required_fields"] = []
    v["invalid_date_formats"] = []
    v["invalid_rows"] = []
    v["warnings"] = [w for w in v["warnings"] if "missing" in w.lower()]

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Demo-data sanitization helpers
# ─────────────────────────────────────────────────────────────────────────────

_GENDERS = ["Male", "Female", "Other", "Prefer not to say"]
_STATUS_DEFAULT = "Working"


def _random_joining_date() -> str:
    """Random date within the last 5 years."""
    today = date.today()
    days_back = random.randint(30, 5 * 365)
    return (today - timedelta(days=days_back)).isoformat()


def _random_birth_date() -> str:
    """Random date producing an age between 25 and 55."""
    today = date.today()
    age_years = random.randint(25, 55)
    days = age_years * 365 + random.randint(0, 364)
    return (today - timedelta(days=days)).isoformat()


def _sanitize_demo_row(emp: dict[str, Any]) -> dict[str, Any]:
    if not emp.get("Date of Joining"):
        emp["Date of Joining"] = _random_joining_date()
    if not emp.get("Date of Birth"):
        emp["Date of Birth"] = _random_birth_date()
    if not emp.get("Gender"):
        emp["Gender"] = random.choice(_GENDERS)
    if not emp.get("Working Status"):
        emp["Working Status"] = _STATUS_DEFAULT
    return emp
